from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ...constants import ELEMENT_ORDER


def _apply_table_borders(ws, first_row: int, last_row: int, first_col: int, last_col: int) -> None:
    thick = Side(style="thick", color="282829")
    thin = Side(style="thin", color="282829")

    for row in ws.iter_rows(min_row=first_row, max_row=last_row, min_col=first_col, max_col=last_col):
        for cell in row:
            r, c = cell.row, cell.column
            cell.border = Border(
                top=thick if r == first_row else thin,
                bottom=thick if r == last_row else thin,
                left=thick if c == first_col else thin,
                right=thick if c == last_col else thin,
            )


class FieldReportGenerator:
    """Genera un Excel de referencia con las reglas de cada campo."""

    HEADERS = [
        "Element", "Field ID", "Label", "Required",
        "Data Type", "Max Length", "Picklist ID", "Visibility", "Business Key",
    ]

    _COLOR_HEADER_BG = "282829"
    _COLOR_HEADER_FONT = "FFFFFF"
    _COLOR_ELEMENT_HDR = "393939"
    _COLOR_REQUIRED = "FBBAB4"
    _COLOR_BIZ_KEY = "F0F4C3"
    _COLOR_ALT_ROW = "F5F5F5"

    def generate(self, metadata: Dict, translated_labels: Dict[str, str], output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Field Reference"

        self._write_header_row(ws)
        last_row = self._write_data_rows(ws, metadata, translated_labels)
        self._apply_column_widths(ws)
        ws.freeze_panes = "A2"

        _apply_table_borders(ws, first_row=1, last_row=last_row, first_col=1, last_col=len(self.HEADERS))

        # Hoja de validacion (si hay issues)
        validation = metadata.get("validation", {})
        issues = validation.get("issues", [])
        if issues:
            self._write_validation_sheet(wb, issues, validation.get("summary", {}))

        wb.save(output_path)
        return output_path

    def _write_header_row(self, ws) -> None:
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color=self._COLOR_HEADER_FONT, size=11, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=self._COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

    def _write_data_rows(self, ws, metadata: Dict, translated_labels: Dict[str, str]) -> int:
        field_catalog: Dict = metadata.get("field_catalog", {})
        rows = self._build_sorted_rows(field_catalog, translated_labels)

        current_element = None
        data_row = 2

        for row_data in rows:
            element = row_data["element"]
            if element != current_element:
                self._write_element_separator(ws, data_row, element)
                current_element = element
                data_row += 1
            self._write_data_row(ws, data_row, row_data, self._resolve_row_fill(row_data, data_row))
            data_row += 1

        return data_row - 1

    def _write_element_separator(self, ws, row: int, element_name: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(self.HEADERS))
        cell = ws.cell(row=row, column=1, value=element_name.upper())
        cell.font = Font(bold=True, color=self._COLOR_HEADER_FONT, size=10, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=self._COLOR_ELEMENT_HDR)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16

    def _write_data_row(self, ws, row: int, data: Dict, fill: PatternFill) -> None:
        values = [
            data["element"],
            data["field_id"],
            data["label"],
            "Yes" if data["required"] else "No",
            data["data_type"],
            data["max_length"] if data["max_length"] is not None else "-",
            data["picklist_id"] or "-",
            data["visibility"],
            "Yes" if data["is_business_key"] else "No",
        ]

        center_cols = {4, 5, 6, 8, 9}

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = fill
            cell.font = Font(size=10, name="Calibri")
            cell.alignment = Alignment(
                horizontal="center" if col_idx in center_cols else "left",
                vertical="center",
                indent=0 if col_idx in center_cols else 1,
            )

        ws.row_dimensions[row].height = 15

    def _build_sorted_rows(self, field_catalog: Dict, translated_labels: Dict[str, str]) -> List[Dict]:
        buckets: Dict[str, List[Dict]] = {}

        for full_id, entry in field_catalog.items():
            element = entry.get("element", "unknown")
            row = {
                "full_id": full_id,
                "element": element,
                "field_id": entry.get("field", full_id),
                "label": translated_labels.get(full_id, entry.get("field", full_id)),
                "required": entry.get("required", False),
                "data_type": entry.get("data_type", "string"),
                "max_length": entry.get("max_length"),
                "picklist_id": entry.get("picklist_id"),
                "visibility": entry.get("visibility", "-"),
                "is_business_key": entry.get("is_business_key", False),
            }
            buckets.setdefault(element, []).append(row)

        for element in buckets:
            buckets[element].sort(key=lambda r: (not r["is_business_key"], r["field_id"]))

        result = []
        for element_id in ELEMENT_ORDER:
            if element_id in buckets:
                result.extend(buckets.pop(element_id))
        for element_id in sorted(buckets):
            result.extend(buckets[element_id])

        return result

    def _resolve_row_fill(self, row_data: Dict, row_idx: int) -> PatternFill:
        if row_data["is_business_key"]:
            return PatternFill("solid", fgColor=self._COLOR_BIZ_KEY)
        if row_data["required"]:
            return PatternFill("solid", fgColor=self._COLOR_REQUIRED)
        if row_idx % 2 == 0:
            return PatternFill("solid", fgColor=self._COLOR_ALT_ROW)
        return PatternFill("solid", fgColor="FFFFFF")

    def _apply_column_widths(self, ws) -> None:
        widths = [18, 30, 35, 10, 12, 12, 22, 12, 14]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Hoja de validacion ───────────────────────────────────────────────

    _VALIDATION_HEADERS = ["Severity", "Code", "Message", "Element", "Field", "Country", "Validator"]
    _SEVERITY_COLORS = {
        "fatal": "D32F2F",
        "error": "F57C00",
        "warning": "FBC02D",
    }

    def _write_validation_sheet(self, wb, issues: List[Dict], summary: Dict) -> None:
        ws = wb.create_sheet("Validation")

        # Header
        for col_idx, header in enumerate(self._VALIDATION_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color=self._COLOR_HEADER_FONT, size=11, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=self._COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        # Summary row
        summary_text = (
            f"Total: {summary.get('total', 0)} | "
            f"Fatal: {summary.get('fatal', 0)} | "
            f"Error: {summary.get('error', 0)} | "
            f"Warning: {summary.get('warning', 0)}"
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(self._VALIDATION_HEADERS))
        cell = ws.cell(row=2, column=1, value=summary_text)
        cell.font = Font(bold=True, size=10, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=self._COLOR_ALT_ROW)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Issue rows
        for row_idx, issue in enumerate(issues, start=3):
            severity = issue.get("severity", "")
            values = [
                severity.upper(),
                issue.get("code", ""),
                issue.get("message", ""),
                issue.get("element_id", "") or "-",
                issue.get("field_id", "") or "-",
                issue.get("country_code", "") or "-",
                issue.get("validator", "") or "-",
            ]

            color = self._SEVERITY_COLORS.get(severity, "FFFFFF")
            fill = PatternFill("solid", fgColor=color)
            severity_fill = fill

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(
                    size=10, name="Calibri",
                    color="FFFFFF" if col_idx == 1 else "000000",
                    bold=col_idx == 1,
                )
                cell.fill = severity_fill if col_idx == 1 else PatternFill("solid", fgColor="FFFFFF")
                cell.alignment = Alignment(
                    horizontal="center" if col_idx in {1, 2} else "left",
                    vertical="center",
                )

        # Column widths
        val_widths = [12, 16, 60, 22, 28, 10, 22]
        for col_idx, width in enumerate(val_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A3"

        last_row = max(3, 2 + len(issues))
        _apply_table_borders(ws, first_row=1, last_row=last_row, first_col=1, last_col=len(self._VALIDATION_HEADERS))
